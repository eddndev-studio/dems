#!/usr/bin/env python3
"""Genera apps/api/crates/seed/src/data.rs desde los Excel oficiales 2026.

Fuentes:
  - prototipos2026_Jurados (1).xlsx  -> hoja "Hoja1" (autoritativa: 21 jurados,
    3 por categoría, todos con correo + categoría).
  - prototipos2026_EtapaFinal.xlsx   -> hoja "Prototipos" (70 prototipos).

NO se incluye la columna "Unidad Académica". El password de cada jurado = su email.

Uso:
  python3 scripts/gen_seed_data.py \
      "/ruta/prototipos2026_Jurados (1).xlsx" \
      "/ruta/prototipos2026_EtapaFinal.xlsx"
"""
import sys
import re
import unicodedata
import openpyxl

OUT = "apps/api/crates/seed/src/data.rs"


def norm(s):
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode().upper()
    return re.sub(r"\s+", " ", s).strip()


def clean(s):
    return re.sub(r"\s+", " ", str(s)).strip()


def rust(s):
    return clean(s).replace("\\", "\\\\").replace('"', '\\"')


# slug por categoría normalizada (cubre las variantes de ambos archivos)
SLUG = {
    "APLICACION EMPRESA": "aplicacion-empresa",
    "APLICACION PARA LA EMPRESA": "aplicacion-empresa",
    "MAQUINARIA Y EQUIPO": "maquinaria-equipo",
    "MAQUINARIA Y EQUIPO PRODUCTIVO": "maquinaria-equipo",
    "PROCESOS QUIMICOS Y BIOLOGICOS": "procesos-quimicos-biologicos",
    "PRODUCTOS PARA LA ENSENANZA": "productos-ensenanza",
    "PRODUCTOS PARA LA SALUD": "productos-salud",
    "SOFTWARE": "desarrollo-software",
    "DESARROLLO DE SOFTWARE": "desarrollo-software",
    "SOLUCIONES DOMESTICAS": "soluciones-domesticas",
}


def slug(cat):
    n = norm(cat)
    if n not in SLUG:
        raise SystemExit(f"categoria sin slug: {cat!r} -> {n!r}")
    return SLUG[n]


def main():
    jurados_xlsx = sys.argv[1] if len(sys.argv) > 1 else \
        "/home/eddndev/Downloads/prototipos2026_Jurados (1).xlsx"
    proto_xlsx = sys.argv[2] if len(sys.argv) > 2 else \
        "/home/eddndev/Downloads/prototipos2026_EtapaFinal.xlsx"

    ws = openpyxl.load_workbook(jurados_xlsx, data_only=True)["Hoja1"]
    jur = []
    for r in list(ws.iter_rows(values_only=True))[1:]:
        nom, ap, am, correo, cat = r[3], r[4], r[5], r[9], r[10]
        if not nom or not correo:
            continue
        nombre = clean(" ".join(x for x in [str(nom), str(ap or ""), str(am or "")] if x.strip()))
        jur.append((clean(correo).lower(), nombre, slug(cat)))
    seen = set()
    jur = [x for x in jur if not (x[0] in seen or seen.add(x[0]))]

    ws = openpyxl.load_workbook(proto_xlsx, data_only=True)["Prototipos"]
    proto = []
    for r in list(ws.iter_rows(values_only=True))[2:]:
        cat, folio, nombre = r[0], r[1], r[2]
        if not folio:
            continue
        proto.append((clean(folio), clean(nombre), slug(cat)))

    out = [
        "// Auto-generado desde los Excel oficiales 2026 (jurados Hoja1 + prototipos EtapaFinal).",
        "// NO se incluye la columna Unidad Académica. Regenerar con scripts/gen_seed_data.py.",
        "",
        "/// (email, nombre completo, slug de categoría). Password del jurado = su email.",
        "pub const JURADOS: &[(&str, &str, &str)] = &[",
    ]
    out += [f'    ("{rust(e)}", "{rust(n)}", "{s}"),' for e, n, s in jur]
    out += [
        "];",
        "",
        "/// (folio, nombre, slug de categoría).",
        "pub const PROTOTIPOS: &[(&str, &str, &str)] = &[",
    ]
    out += [f'    ("{rust(f)}", "{rust(n)}", "{s}"),' for f, n, s in proto]
    out += ["];"]

    with open(OUT, "w", encoding="utf-8") as fh:
        fh.write("\n".join(out) + "\n")
    print(f"escrito {OUT}: jurados={len(jur)} prototipos={len(proto)}")


if __name__ == "__main__":
    main()
